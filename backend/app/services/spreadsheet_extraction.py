from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


CONFIG_SHEET_KEYWORDS = {
    "config",
    "configuration",
    "lookup",
    "mapping",
    "parameter",
    "profile",
    "setup",
    "value",
    "values",
}


def serialize_cell_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()

    return value


def cell_has_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def get_used_bounds(worksheet: Any) -> tuple[int, int, int, int]:
    try:
        dimension = worksheet.calculate_dimension()
    except ValueError:
        dimension = worksheet.calculate_dimension(force=True)

    if dimension == "A1:A1" and not cell_has_value(worksheet["A1"].value):
        return 1, 1, 0, 0

    return range_boundaries(dimension)


def detect_header_rows(rows: list[dict[str, Any]]) -> list[int]:
    header_rows: list[int] = []

    for row in rows[:5]:
        values = row["values"]
        non_empty_values = [value for value in values if cell_has_value(value)]
        string_values = [value for value in non_empty_values if isinstance(value, str)]

        if (
            len(non_empty_values) >= 2
            and len(string_values) >= len(non_empty_values) / 2
        ):
            header_rows.append(row["row_number"])
            break

    return header_rows


def is_likely_config_sheet(
    sheet_name: str,
    non_empty_row_count: int,
    max_column: int,
) -> bool:
    normalized_name = sheet_name.lower()

    if any(keyword in normalized_name for keyword in CONFIG_SHEET_KEYWORDS):
        return non_empty_row_count > 0

    return non_empty_row_count >= 2 and max_column >= 2


def extract_visible_sheet(
    worksheet: Any,
    max_rows_per_sheet: int,
) -> dict[str, Any]:
    min_column, min_row, max_column, max_row = get_used_bounds(worksheet)
    extracted_rows: list[dict[str, Any]] = []
    non_empty_row_count = 0

    if max_row == 0 or max_column == 0:
        return {
            "name": worksheet.title,
            "max_row": 0,
            "max_column": 0,
            "non_empty_row_count": 0,
            "detected_header_rows": [],
            "is_likely_config_sheet": False,
            "rows": [],
        }

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_column,
            max_col=max_column,
        ),
        start=min_row,
    ):
        values = [serialize_cell_value(cell.value) for cell in row]

        if not any(cell_has_value(value) for value in values):
            continue

        non_empty_row_count += 1

        if len(extracted_rows) < max_rows_per_sheet:
            extracted_rows.append(
                {
                    "row_number": row_number,
                    "values": values,
                }
            )

    detected_header_rows = detect_header_rows(extracted_rows)

    return {
        "name": worksheet.title,
        "max_row": max_row,
        "max_column": max_column,
        "non_empty_row_count": non_empty_row_count,
        "detected_header_rows": detected_header_rows,
        "is_likely_config_sheet": is_likely_config_sheet(
            worksheet.title,
            non_empty_row_count,
            max_column,
        ),
        "rows": extracted_rows,
    }


def render_sheet_text(sheet: dict[str, Any]) -> str:
    lines = [
        f"Sheet: {sheet['name']}",
        f"Used range: {sheet['max_row']} rows x {sheet['max_column']} columns",
        f"Non-empty rows: {sheet['non_empty_row_count']}",
        f"Likely config sheet: {'yes' if sheet['is_likely_config_sheet'] else 'no'}",
    ]

    if sheet["detected_header_rows"]:
        header_rows = ", ".join(str(row) for row in sheet["detected_header_rows"])
        lines.append(f"Detected header rows: {header_rows}")

    if sheet["rows"]:
        lines.append("Rows:")

        for row in sheet["rows"]:
            rendered_values = [
                "" if value is None else str(value) for value in row["values"]
            ]
            lines.append(f"{row['row_number']}: " + " | ".join(rendered_values))

    return "\n".join(lines)


def extract_spreadsheet(
    file_path: Path,
    max_rows_per_sheet: int,
) -> dict[str, Any]:
    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=False,
        keep_vba=file_path.suffix.lower() == ".xlsm",
    )

    try:
        sheets = [
            extract_visible_sheet(worksheet, max_rows_per_sheet)
            for worksheet in workbook.worksheets
            if worksheet.sheet_state == "visible"
        ]
        metadata = {
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": list(workbook.sheetnames),
        }

        return {
            "text_content": "\n\n".join(render_sheet_text(sheet) for sheet in sheets),
            "json_content": {
                "workbook": metadata,
                "sheets": sheets,
            },
        }
    finally:
        vba_archive = getattr(workbook, "vba_archive", None)
        if vba_archive is not None:
            vba_archive.close()

        workbook.close()
